import ast
import os
import time
from io import BytesIO
from urllib.request import urlopen
from zipfile import ZipFile
import secrets
import requests

# debug onl yin developmwnt
from django.views import debug
import sys

from .imagekit_media import ImageMediaLibrary

from cryptography.hazmat.primitives.asymmetric import padding
import base64
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import serialization, padding
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes

import openpyxl
import pandas as pd
from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import Group
from django.core.cache import cache

from django.core.exceptions import (
    MultipleObjectsReturned,
    ObjectDoesNotExist,
    PermissionDenied,
    SuspiciousOperation,
    ValidationError,
)
from .custom_exceptions import (
    SessionValuesNotFoundError,
    SimilarItemHeadingError,
    SimilarItemHeadingDataError,
    HeaderDataNotFoundError,
    TableNotFoundError,
    ImageMediaStorageError,
)
from django.core.files.base import ContentFile
from django.db import (
    DatabaseError,
    DataError,
    IntegrityError,
    OperationalError,
    transaction,
)
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.shortcuts import redirect, render
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils.encoding import smart_str
from django.views import View
from PIL import Image, ImageColor, ImageDraw, ImageFont
from reportlab.lib.colors import Color
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas


import ast
import os
import time
from io import BytesIO
from urllib.request import urlopen
from zipfile import ZipFile

import openpyxl
import pandas as pd
from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import Group
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.db import DatabaseError, DataError, IntegrityError, OperationalError
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import redirect, render
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils.encoding import smart_str
from PIL import Image, ImageColor, ImageDraw, ImageFont
from reportlab.lib.colors import Color
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from .forms import (
    ExcelForm,
    ExportForm,
    ImageForm,
    ItemForm,
    LoginForm,
    NameSignUpForm,
    SignUpForm,
)
from .models import CustomUser, DataItemSetModel, ImageModel


def check_time(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        elapsed_time = end_time - start_time
        print(
            f"\nTIMER: '{func.__name__}' took {elapsed_time:.6f} seconds to execute.\n"
        )
        return result

    return wrapper


def page_renderer(func):
    """
    Decorator that handles various exceptions and renders the page with appropriate error information.

    Raises: ObjectDoesNotExist, ValidationError, SuspiciousOperation, PermissionDenied, Http404, IntegrityError, DataError, DatabaseError, OperationalError
    """

    def wrapper(
        *args, **kwargs
    ):  # TODO: Handle Errors for instance.save() and instance.delete()
        try:
            success = False

            instance = args[0]
            func(*args, **kwargs)

            success = True

        except ObjectDoesNotExist as e:
            if hasattr(e, "form_name") and e.form_name:
                instance.context[f"{e.form_name}_errors"]["error"] = None
            else:
                instance.context["db_error"] = f"Object does not exist: {e.__str__()}"

            # used for debugging
            return debug.technical_500_response(instance.request, *sys.exc_info())

        except MultipleObjectsReturned as e:
            if hasattr(e, "form_name") and e.form_name:
                instance.context[f"{e.form_name}_errors"]["error"] = None
            else:
                instance.context["db_error"] = f"Multiple Objects Found: {e.__str__()}"

            # used for debugging
            return debug.technical_500_response(instance.request, *sys.exc_info())

        except ValidationError as e:
            print(e.__str__())

            if hasattr(e, "form_name") and e.form_name:
                instance.context[f"{e.form_name}_errors"]["error"] = None
            else:
                instance.context["db_error"] = f"Validation Error: {e.__str__()}"

            # used for debugging
            return debug.technical_500_response(instance.request, *sys.exc_info())

        except SuspiciousOperation as e:
            print(e.__str__())

            if hasattr(e, "form_name") and e.form_name:
                instance.context[f"{e.form_name}_errors"]["error"] = None
            else:
                instance.context[
                    "db_error"
                ] = f"A suspicious operation was detected: {e.__str__()}"

            # used for debugging
            return debug.technical_500_response(instance.request, *sys.exc_info())

        except PermissionDenied:
            print(
                "Permission Denied: You do not have permission to perform this action."
            )

            if hasattr(e, "form_name") and e.form_name:
                instance.context[f"{e.form_name}_errors"]["error"] = None
            else:
                instance.context[
                    "db_error"
                ] = "Permission Denied: You do not have permission to perform this action."

            # used for debugging
            return debug.technical_500_response(instance.request, *sys.exc_info())

        except ConnectionError as e:  # add any cache errors
            print(e.__str__())

            if hasattr(e, "form_name") and e.form_name:
                instance.context[f"{e.form_name}_errors"]["error"] = None
            else:
                instance.context["db_error"] = f"Connection Error: {print(e.__str__())}"

            # used for debugging
            return debug.technical_500_response(instance.request, *sys.exc_info())

        except TypeError as e:
            print(e.__str__())

            if hasattr(e, "form_name") and e.form_name:
                instance.context[f"{e.form_name}_errors"]["error"] = None
            else:
                instance.context["db_error"] = f"Invalid Type: {print(e.__str__())}"

            # used for debugging
            return debug.technical_500_response(instance.request, *sys.exc_info())

        except Http404:
            print("HTTP 404: The requested resource was not found.")

            if hasattr(e, "form_name") and e.form_name:
                instance.context[f"{e.form_name}_errors"]["error"] = None
            else:
                instance.context["db_error"] = "The requested resource was not found."

            # used for debugging
            return debug.technical_500_response(instance.request, *sys.exc_info())

        except IntegrityError as e:
            print(e.__str__())

            if hasattr(e, "form_name") and e.form_name:
                instance.context[f"{e.form_name}_errors"]["error"] = None
            else:
                instance.context[
                    "db_error"
                ] = f"Integrity Error: Unique Constraint Violated ({e.__str__()})"

            # used for debugging
            return debug.technical_500_response(instance.request, *sys.exc_info())

        except DataError as e:
            print(e.__str__())

            if hasattr(e, "form_name") and e.form_name:
                instance.context[f"{e.form_name}_errors"]["error"] = None
            else:
                instance.context[
                    "db_error"
                ] = f"Data Error: Invalid data types or lengths. ({e.__str__()})"

            # used for debugging
            return debug.technical_500_response(instance.request, *sys.exc_info())

        except DatabaseError as e:
            print(e.__str__())

            if hasattr(e, "form_name") and e.form_name:
                instance.context[f"{e.form_name}_errors"]["error"] = None
            else:
                instance.context["db_error"] = f"Database Error: ({e.__str__()})"

            # used for debugging
            return debug.technical_500_response(instance.request, *sys.exc_info())

        except OperationalError as e:
            if hasattr(e, "form_name") and e.form_name:
                instance.context[f"{e.form_name}_errors"]["error"] = None
            else:
                instance.context[
                    "db_error"
                ] = f"Operational Error: Connection problem or timeout. ({e.__str__()})"

            # used for debugging
            return debug.technical_500_response(instance.request, *sys.exc_info())

        except Exception as e:
            print(e.__str__())

            # Generic Error
            if hasattr(e, "form_name") and e.form_name:
                instance.context[f"{e.form_name}_errors"]["error"] = None
            else:
                instance.context[
                    "db_error"
                ] = f"An unexpected error occurred: {e.__str__()}"

            # used for debugging
            return debug.technical_500_response(instance.request, *sys.exc_info())

        finally:
            if success:
                instance.context["db_error"] = None
                return render(
                    instance.request, instance.home_template, instance.context
                )

            else:
                pass

    return wrapper


class IndexView(View):
    # Reload the Cache and update the session variables accordingly
    @check_time
    def reload_cache(self, headers, header_items):
        """Update the Cache for Header items and Header Data. It also Updates the Respective Session Variables"""

        try:
            if headers:
                inspector_header = list(
                    DataItemSetModel.objects.filter(
                        user_code=self.request.session.get("user_code")
                    ).values_list("item_set_heading", flat=True)
                )

            if header_items is not None:
                if header_items[0] == "__all__":
                    if headers:
                        header_items = inspector_header
                    elif not headers and cache.get(self.cache_key_header) is not None:
                        header_items = cache.get(self.cache_key_header)
                    else:
                        raise ObjectDoesNotExist()

                inspector_data_buffer = {}
                for item in header_items:
                    inspector_data = ast.literal_eval(
                        DataItemSetModel.objects.get(
                            item_set_heading=item,
                            user_code=self.request.session.get("user_code"),
                        ).item_set
                    )
                    inspector_data_buffer[item] = inspector_data

                for item in header_items:
                    cache.set(item, inspector_data_buffer[item])

                if self.request.session.get("current_header", "") != "":
                    self.context["inspector_data"] = cache.get(
                        self.request.session.get("current_header")
                    )

            if headers:
                cache.set(self.cache_key_header, inspector_header)
                self.request.session["inspector_header"] = inspector_header

        except KeyError as e:
            reload_cache(headers=True, header_items=None)
            reload_cache(headers=False, header_items=header_items)

        except ObjectDoesNotExist as e:
            print("reload_cache: " + "No Item Data in DB")
            return False

        except Exception as e:
            print("reload_cache: " + e.__str__())
            raise

    # Initliaze the Context Dictionary
    def init_context(self):
        """Initializes the context dictionary which stores all the metadata about the user"""

        self.context["new_user"] = True
        self.context["set_cookie"] = False
        self.context["cookie_key"] = None
        self.context["cookie_data"] = None
        self.request.session["cookie_is_set"] = None

        # keep unchanged if is True or False, and set it to None otherwise
        if self.request.session.get("cookie_consent") not in [True, False]:
            self.request.session["cookie_consent"] = None

        if self.context.get("db_error") is None:
            self.context["db_error"] = None

        self.context["login_form"] = None
        self.context["login_form_errors"] = {"has_error": False, "error": None}

        self.context["name_signup_form_errors"] = {"has_error": False, "error": None}

        self.context["excel_form"] = None
        self.context["excel_form_errors"] = {"has_error": False, "error": None}
        self.context["excel_file_status"] = None

        self.context["item_form"] = None
        self.context["item_form_errors"] = {"has_error": False, "error": None}

        self.context["image_form"] = None
        self.context["image_form_errors"] = {"has_error": False, "error": None}
        self.context["image_status"] = None
        self.context["image_url"] = None

        self.context["export_form"] = None
        self.context["export_form_errors"] = {"has_error": False, "error": None}

        # TODO: REMOVE
        # if cache.get(self.cache_key_header) is not None:
        #     self.context["inspector_header"] = cache.get(self.cache_key_header)

        # if self.request.session.get("current_header", "") != "":
        #     self.context["inspector_data"] = cache.get(
        #         self.request.session.get("current_header")
        #     )

        # self.context["inspector_data"] = cache.get(
        #     self.request.session.get("current_header")
        # )

    # Setup the Forms
    def verify_form_data(self):
        """Setup the Form Values and Session Variables"""

        self.context["login_form"] = LoginForm()

        self.context[
            "excel_form"
        ] = ExcelForm()  # IDEA: add excel_file name as the data

        if (
            self.request.session.get("current_header", None) is not None
            and self.request.session.get("user_code", None) is not None
        ):
            try:
                instance = DataItemSetModel.objects.get(
                    user_code=self.request.session.get("user_code"),
                    item_set_heading=self.request.session.get("current_header"),
                )

                self.context["item_form"] = ItemForm(
                    initial={
                        "item_heading": self.request.session.get("current_header"),
                        "color": instance.color,
                    },
                    instance=instance,
                )
                self.context["format_reverse"] = True
            except ObjectDoesNotExist:
                self.context["item_form"] = ItemForm()

            except Exception as e:
                print("init_context" + e.__str__())
                raise

        else:
            if self.request.session.get("current_header", None) is not None:
                if DataItemSetModel.objects.filter(
                    user_code=self.request.session.get("user_code")
                ).exists():
                    latest_instance = (
                        DataItemSetModel.objects.filter(
                            user_code=self.request.session.get("user_code")
                        )
                        .order_by("-created")
                        .first()
                    )
                    if latest_instance:
                        self.request.session[
                            "current_header"
                        ] = latest_instance.item_set_heading

            self.context["item_form"] = ItemForm()

        self.context["image_form"] = ImageForm()

        self.context["export_form"] = ExportForm()

    # Update the user type
    def verify_user_type(self):
        """Updates the user type data in the context dictionary."""

        if self.request.session.get("user_code", "") != "":
            self.context["new_user"] = False

            self.cache_key_header = (
                f'{self.request.session.get("user_code")}-db_cache_headers'
            )

            if self.request.session.get(
                "cookie_is_set"
            ) is not True and self.request.session.get("cookie_consent", False):
                self.context["set_cookie"] = True
                self.request.session["cookie_is_set"] = True
                self.context["cookie_key"] = self.cookie_key
                self.context[
                    "cookie_data"
                ] = self.encrypted_cookie_data()  # TODO: check whole functionality
                self.context["cookie_data_temp"] = self.decrypt_cookie_data(
                    self.context.get("cookie_data")
                )  # TODO: Remove

            elif self.request.session.get("cookie_is_set"):
                self.context["set_cookie"] = False
                self.context["cookie_key"] = None
                self.context["cookie_data"] = None
                self.request.session["cookie_is_set"] = False

        # Validates the Cookie Save Consent
        if self.request.POST.get("allow_cookies", None) is not None:
            self.request.session["cookie_consent"] = bool(
                self.request.POST.get("allow_cookies")
            )

        # Checks if the Cookie Stored is Valid
        if (
            self.request.session.get("cookie_is_set") is False
            and self.request.COOKIES.get(self.cookie_key) is not None
        ):
            if (
                self.decrypt_cookie_data(self.request.COOKIES.get(self.cookie_key))
                is not False
            ):
                if CustomUser.objects.filter(
                    unique_code=self.request.COOKIES.get(self.cookie_key)
                ).exists():
                    self.request.session["user_code"] = self.request.COOKIES.get(
                        self.cookie_key
                    )

                    self.context["new_user"] = False

                else:
                    self.context["new_user"] = True
                    self.context["name_signup_form"] = NameSignUpForm()
                    self.request.session["cookie_is_set"] = None

            else:
                self.context["new_user"] = True
                self.context["name_signup_form"] = NameSignUpForm()
                self.request.session["cookie_is_set"] = None

        else:
            self.context["name_signup_form"] = NameSignUpForm()

        # check if the user is Logged in or has a Session Started
        if self.request.user.is_authenticated and isinstance(
            self.request.user, CustomUser
        ):
            self.request.session["is_verified"] = True

        else:
            self.request.session["is_verified"] = False

        # reload cache and display headers if possible
        if (
            len(cache.get(self.cache_key_header, [])) == 0
            and self.request.session.get("user_code", "") != ""
        ):
            if self.reload_cache(
                headers=True,
                header_items=["__all__"],
            ):
                self.context["inspector_data"] = cache.get(
                    self.request.session.get("current_header")
                )
        else:
            if self.request.session.get("current_header") is not None:
                self.context["inspector_data"] = cache.get(
                    self.request.session.get("current_header")
                )

    # Render Preview Image
    @check_time
    def render_preview_url(self):  # TODO: Verify
        """Renders a Single Image Using the First Set of Items for Preview"""

        if self.request.session.get("image_url", "") != "":
            if self.request.session.get("user_code", "") != "":
                try:
                    data_items = DataItemSetModel.objects.filter(
                        user_code=self.request.session.get("user_code")
                    )
                    image_model = ImageModel.objects.get(
                        user__unique_code=self.request.session.get("user_code")
                    )

                    response = requests.get(imagekit_url)

                    if response.status_code == 200:
                        image = Image.open(BytesIO(response.content))

                    else:
                        raise ImageMediaStorageError(
                            "Image was Not Found in the Media Storage"
                        )

                    image = image.convert("RGBA")

                    draw = ImageDraw.Draw(image)

                    for data_item in data_items:
                        font_path = DataItemSetModel.search_font(data_item.font_name)
                        font = ImageFont.truetype(
                            font_path, data_item.font_size + 50
                        )  # FIXME: remove added increment

                        # take # into consideration and
                        # Extract RGB values and transparency from data_item.color
                        rgb_values = tuple(
                            int(data_item.color[i : i + 2], 16) for i in (1, 3, 5)
                        )

                        # IDK how this works
                        transparency = int(data_item.color[-2:], 16) / 255.0

                        text_position = (data_item.position_x, data_item.position_y)

                        draw.text(
                            text_position,
                            str(ast.literal_eval(data_item.item_set)[0]),
                            font=font,
                            fill=rgb_values + (int(transparency * 255),),
                        )

                    _, extension = os.path.splitext(image_model.image_file_name)
                    extension = extension[1:].upper()

                    rendered_image = BytesIO()
                    image.save(rendered_image, format=extension)

                    # Save the modified image to the preview storage

                    preview_image_storage = ImageMediaLibrary("preview")

                    image_upload_result = preview_image_storage.upload_image(
                        image_file=rendered_image,
                        tags=[
                            self.request.session.get("user_code"),
                        ],
                        overwrite_status=True,
                    )

                    if image_upload_result:
                        if image_model.preview_image_url != "":
                            if not preview_image_storage.delete_image(
                                image_url=image_model.preview_image_url,
                                image_type="preview",
                                user_code=self.request.session.get("user_code"),
                            ):
                                raise ImageMediaStorageError(
                                    "The Previous Preview Image Could not be Deleted"
                                )

                        image_model.preview_image_url = image_upload_result.url

                    else:
                        raise ImageMediaStorageError(
                            "Image Could not be Uploaded to the Media Storage"
                        )

                    image_model.save()

                    self.request.session["preview_url"] = image_model.preview_image_url

                except Exception as e:
                    print("render_preview_url" + e.__str__())
                    raise

            else:
                raise SessionValuesNotFoundError("user_code Not Available")

    ############
    def render_and_zip_images(user_code, output_format="png"):
        # Fetch the background image for the user_code
        image_model = ImageModel.objects.get(user__unique_code=user_code)

        # Create a folder with user_code to store temporary files
        temp_folder = os.path.join(settings.MEDIA_ROOT, user_code)
        os.makedirs(temp_folder, exist_ok=True)

        # Fetch data items for the user_code
        data_items = DataItemSetModel.objects.filter(user_code=user_code)

        # Group data items by their item_set index position
        grouped_data_items = {}
        for data_item in data_items:
            index_position = data_item.item_set.index
            if index_position not in grouped_data_items:
                grouped_data_items[index_position] = []
            grouped_data_items[index_position].append(data_item)

        # Initialize a list to store file paths
        file_paths = []

        for index_position, data_items_at_index in grouped_data_items.items():
            # Set the text color with alpha
            text_color = hex_to_rgb_with_alpha(data_items_at_index[0].color)

            # Render text on the image
            image_path = render_text_on_image(data_items_at_index, image_model)

            # Save the image based on the output format
            if output_format == "png":
                output_path = os.path.join(
                    temp_folder, f"output_index_{index_position}.png"
                )
                save_image_as_png(image_path, output_path)
            elif output_format == "jpeg":
                output_path = os.path.join(
                    temp_folder, f"output_index_{index_position}.jpeg"
                )
                save_image_as_jpeg(image_path, output_path)
            elif output_format == "pdf":
                output_path = os.path.join(
                    temp_folder, f"output_index_{index_position}.pdf"
                )
                create_pdf([image_path], output_path)
            else:
                raise ValueError(
                    "Invalid output format. Supported formats are 'png', 'jpeg', and 'pdf'."
                )

            file_paths.append(output_path)

        # Zip all the files
        zip_path = os.path.join(settings.MEDIA_ROOT, f"{user_code}_output.zip")
        zip_files(zip_path, temp_folder)

        # Clean up: Delete temporary folder
        delete_temp_folder(temp_folder)

        return zip_path

    def hex_to_rgb_with_alpha(hex_color):
        # Decode hex color with alpha (##RRGGBBAA)
        return Color(hex_color).rgb

    def render_text_on_image(data_item, image_model):
        # Open the image
        image = Image.open(image_model.image.path)

        # Create a drawing object
        draw = ImageDraw.Draw(image)

        # Load the specified font
        font_path = DataItemSetModel.search_font(data_item.font_name)
        font = ImageFont.truetype(font_path, data_item.font_size)

        # Set the text color with alpha
        text_color = Color(data_item.color)

        # Calculate text position
        text_position = (data_item.position_x, data_item.position_y)

        # Draw the text on the image
        draw.text(text_position, data_item.item_set, font=font, fill=text_color)

        # Save the modified image to a temporary location
        temp_image_path = os.path.join(
            settings.MEDIA_ROOT, f"{get_random_string(8)}.png"
        )
        image.save(temp_image_path)

        return temp_image_path

    def save_image_as_png(input_path, output_path):
        # Open the image
        image = Image.open(input_path)

        # Save the image as PNG
        image.save(output_path, format="PNG")

    def save_image_as_jpeg(input_path, output_path):
        # Open the image
        image = Image.open(input_path)

        # Save the image as JPEG
        image.save(output_path, format="JPEG")

    def create_pdf(image_paths, output_path):
        # Create a PDF with all the images
        packet = BytesIO()
        pdf = canvas.Canvas(packet, pagesize=letter)

        for image_path in image_paths:
            pdf.drawInlineImage(image_path, 0, 0)

        pdf.save()

        # Move to the beginning of the BytesIO buffer
        packet.seek(0)

        # Create a new PDF with the BytesIO content
        with open(output_path, "wb") as pdf_output:
            pdf_output.write(packet.read())

    def zip_files(zip_path, folder_path):
        # Zip all the files in the folder
        with ZipFile(zip_path, "w") as zip_file:
            for foldername, subfolders, filenames in os.walk(folder_path):
                for filename in filenames:
                    file_path = os.path.join(foldername, filename)
                    arcname = os.path.relpath(file_path, folder_path)
                    zip_file.write(file_path, arcname)

    def delete_temp_folder(temp_folder):
        # Delete the temporary folder and its contents
        for file_name in os.listdir(temp_folder):
            file_path = os.path.join(temp_folder, file_name)
            if os.path.isfile(file_path):
                os.unlink(file_path)
            else:
                os.rmdir(file_path)

        # Delete the empty folder
        os.rmdir(temp_folder)

    # Excel File Helper Methods

    # Returns the Table Size and Headings
    def find_table_properties(self, excel_file):
        wb = openpyxl.load_workbook(excel_file)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            for row in sheet.iter_rows(
                min_row=1, max_col=sheet.max_column, max_row=sheet.max_row
            ):
                for cell in row:
                    if cell.value is not None and str(cell.value).strip() != "":
                        start_row = cell.row
                        start_col = cell.column

                        table_length = 0
                        for i in range(start_row, sheet.max_row + 1):
                            if sheet.cell(row=i, column=start_col).value is not None:
                                table_length += 1
                            else:
                                break

                        table_headings = []
                        for i in range(start_col, sheet.max_column + 1):
                            if sheet.cell(row=start_row, column=i).value is not None:
                                table_headings.append(
                                    sheet.cell(row=start_row, column=i).value
                                )
                            else:
                                break

                        return table_headings, start_row, start_col, table_length

        return None

    # Converts Excel File into a Pandas Dataframe
    def excel_to_dataframe(self, excel_file):
        table_info = self.find_table_properties(excel_file)

        if table_info is not None:
            table_headings, start_row, start_col, table_length = table_info

            df = pd.read_excel(
                excel_file,
                header=None,
                skiprows=start_row,
                names=table_headings,
                nrows=table_length,
            )
            df.reset_index(drop=True, inplace=True)

            # Check for duplicate column names
            if len(set(df.columns)) != len(df.columns):
                raise SimilarItemHeadingError(
                    "Duplicate column names found in the DataFrame."
                )

            return df

        else:
            raise TableNotFoundError("The Table was Not Found in this Excel Sheet")

    @check_time
    def store_excel_to_model(self, excel_file):
        """"""

        if self.request.session.get("user_code", "") != "":
            df = self.excel_to_dataframe(excel_file)

            instance_item_headings = DataItemSetModel.objects.filter(
                user_code=self.request.session.get("user_code"),
            ).values_list("item_set_heading", flat=True)

            with transaction.atomic():
                for heading in df.columns:
                    final_heading = str(heading).capitalize().replace(" ", "_")

                    if final_heading not in instance_item_headings:
                        instance = DataItemSetModel()

                        instance.item_set_heading = final_heading
                        instance.item_set = str(df[heading].tolist())
                        instance.user_code = self.request.session.get("user_code")

                        instance.save()

                    else:
                        raise SimilarItemHeadingDataError(
                            f'"{final_heading}" already Exists and Contains Data',
                            old_data=DataItemSetModel.objects.get(
                                user_code=self.request.session.get("user_code"),
                                item_set_heading=final_heading,
                            ).item_set,
                            new_data=str(df[heading].tolist()),
                        )

        else:
            raise SessionValuesNotFoundError("user_code Not Available")

    # Encryption for Cookie
    def decrypt_cookie_data(self, encrypted_data_with_iv):
        try:
            decryption_key = settings.SECRET_KEY[:32].encode(
                "utf-8"
            )  # Ensure it's 32 bytes and encoded
            iv_size = 16  # Size of the IV
            iv = base64.b64decode(
                encrypted_data_with_iv[:iv_size]
            )  # Extract IV from the ciphertext
            ciphertext = base64.b64decode(
                encrypted_data_with_iv[iv_size:]
            )  # Extract ciphertext

            cipher = Cipher(
                algorithms.AES(decryption_key), modes.CFB(iv), backend=default_backend()
            )
            decryptor = cipher.decryptor()
            decrypted_data = decryptor.update(ciphertext) + decryptor.finalize()
            unpadder = padding.PKCS7(128).unpadder()
            unpadded_data = unpadder.update(decrypted_data) + unpadder.finalize()
            return unpadded_data.decode("utf-8")
        except Exception as e:
            print(e.__str__())
            return None

    def encrypted_cookie_data(self):
        try:
            encryption_key = settings.SECRET_KEY[:32].encode(
                "utf-8"
            )  # Ensure it's 32 bytes and encoded
            data = self.request.session.get("user_code")

            # Generate a random IV
            iv = os.urandom(16)

            # Padding the data before encryption
            padder = padding.PKCS7(128).padder()
            padded_data = padder.update(data.encode("utf-8")) + padder.finalize()

            cipher = Cipher(
                algorithms.AES(encryption_key), modes.CFB(iv), backend=default_backend()
            )
            encryptor = cipher.encryptor()
            encrypted_data = base64.b64encode(
                iv + encryptor.update(padded_data) + encryptor.finalize()
            ).decode("utf-8")
            return encrypted_data
        except Exception as e:
            print(e.__str__())
            return None

    home_template = "index.html"
    cookie_key = "autoficate-key"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.context = {}
        self.request = {}
        self.cache_key_header = ""

    @page_renderer
    def dispatch(self, request, *args, **kwargs):
        try:
            self.request = request

            if self.context is not None:
                self.init_context()

            self.verify_form_data()
            self.verify_user_type()

        except Exception:
            raise

        return super().dispatch(request, *args, **kwargs)

    ## Main Request Handlers ##
    @page_renderer
    def get(self, request, *args, **kwargs):
        pass

    @page_renderer
    def post(self, request, *args, **kwargs):
        start_time = time.time()

        # Check and Set User Information Status
        if (
            self.request.POST.get("submit_name_signup") is not None
            and self.request.POST.get("submit_name_signup") == "name_signup"
        ):
            name_signup_form = NameSignUpForm(self.request.POST)
            if name_signup_form.is_valid():
                instance = CustomUser()

                instance.user_email = self.request.POST.get("user_email", "")

                if self.request.POST.get("user_email", "").strip() != "":
                    instance.user_email = (
                        f"{instance.user_email}.code_placeholder.unregistered"
                    )

                else:
                    instance.user_email = ".code_placeholder.unregistered"

                instance.first_name = self.request.POST.get("first_name")
                instance.last_name = self.request.POST.get("last_name")

                instance.save()

                group_instance = Group.objects.get(name="Users")
                user_instance = CustomUser.objects.get(unique_code=instance.unique_code)
                user_instance.groups.set([group_instance])

                user_instance.user_email = user_instance.user_email.replace(
                    "code_placeholder", user_instance.unique_code
                )

                user_instance.save()

                self.request.session["user_name"] = user_instance.username
                self.request.session["user_code"] = user_instance.unique_code

                self.verify_user_type()

            else:
                self.context["name_signup_form_errors"]["has_errors"] = True
                self.context["name_signup_form_errors"][
                    "error"
                ] = name_signup_form.errors

        # Login Form
        elif (
            self.request.POST.get("login") is not None
            and self.request.POST.get("login") == "login"
        ):
            login_form = LoginForm(self.request.POST)

            if login_form.is_valid():
                user_email = login_form.cleaned_data["user_email"]
                password = login_form.cleaned_data["password"]

                user = authenticate(request, user_email=user_email, password=password)

                if user is None:
                    self.context["login_form_errors"]["has_error"] = True
                    self.context["login_form_errors"]["error"] = "Invalid Credentials"

                else:
                    if not user.is_active:
                        # Handle inactive user error
                        self.context["login_form_errors"]["has_error"] = True
                        self.context["login_form_errors"][
                            "error"
                        ] = "Your account is inactive. Please contact support."
                    else:
                        login(request, user)

                        self.request.session["user_name"] = user.username
                        self.request.session["user_code"] = user.unique_code

                        self.verify_user_type()

            else:
                self.context["login_form_errors"]["has_error"] = True
                self.context["login_form_errors"]["error"] = login_form.errors

        # Adds New Blank Data Item in Model
        elif (
            self.request.POST.get("submit_add") is not None
            and self.request.POST.get("submit_add") == "add_blank_item_heading"
        ):
            if self.request.session.get("user_code", "") != "":
                filter_instance = DataItemSetModel.objects.filter(
                    user_code=self.request.session.get("user_code"),
                    item_set_heading="",
                )

                if not filter_instance.exists():
                    instance = DataItemSetModel()

                    instance.user_code = self.request.session.get("user_code")

                    instance.save()

                    self.context["item_form"] = ItemForm(
                        instance=instance,
                    )
                else:
                    self.context["item_form"] = ItemForm(
                        instance=filter_instance.first(),
                    )
            else:
                raise SessionValuesNotFoundError("user_code Not Available")

        # Updates the Data of Existing Blank Item in Model
        elif (
            self.request.POST.get("submit_update") is not None
            and self.request.POST.get("submit_update") == "update_item_heading"
        ):
            item_form = ItemForm(self.request.POST)

            if item_form.is_valid() and self.request.session.get("user_code", "") != "":
                try:
                    # because the Empty header Data is stored as "[]"
                    if len(cache.get(self.cache_key_header, [])) == 0:
                        instance = DataItemSetModel.objects.filter(
                            item_set_heading="",
                            user_code=self.request.session.get("user_code"),
                        )

                        if not instance.exists():
                            instance = DataItemSetModel()
                            instance.user_code = self.request.session.get("user_code")

                    else:
                        instance = DataItemSetModel.objects.get(
                            item_set_heading=self.request.session.get("current_header"),
                            user_code=self.request.session.get("user_code"),
                        )

                except ObjectDoesNotExist as e:
                    print("update_item" + e.__str__())

                    try:
                        if self.request.session.get("current_header", "") != "":
                            instance = DataItemSetModel.objects.get(
                                item_set_heading=self.request.session.get(
                                    "current_header"
                                ),
                                user_code=self.request.session.get("user_code"),
                            )

                        elif item_form.cleaned_data.get("item_heading", "") != "":
                            instance = DataItemSetModel.objects.get(
                                item_set_heading=item_form.cleaned_data.get(
                                    "item_heading"
                                ),
                                user_code=self.request.session.get("user_code"),
                            )

                            self.request.session[
                                "current_header"
                            ] = item_form.cleaned_data.get("item_heading")

                        else:
                            print("update_item" + "Everything Empty")
                            raise HeaderDataNotFoundError(
                                "The Current Header is Missing"
                            )

                    except ObjectDoesNotExist as e:
                        print("update_item" + "Everything Empty")
                        raise HeaderDataNotFoundError("The Current Header is Missing")
                    """ if DataItemSetModel.objects.filter(
                        user_code=self.request.session.get("user_code"),
                        item_set_heading=item_form.cleaned_data.get("item_heading"),
                    ).exists():
                        raise SimilarItemHeadingError(
                            "The Entered Item Heading already Exists in the Database"
                        ) """
                with transaction.atomic():
                    instance.item_set_heading = item_form.cleaned_data.get(
                        "item_heading"
                    )
                    instance.position_x = item_form.cleaned_data.get("position_x")
                    instance.position_y = item_form.cleaned_data.get("position_y")
                    instance.font_size = item_form.cleaned_data.get("font_size")
                    instance.color = item_form.cleaned_data.get("color")
                    instance.font_name = item_form.cleaned_data.get(
                        "font_select"
                    )  # TODO: match font name in form and actual

                    instance.save()

                # to ensure that session.current_header reflects the latest header name
                self.request.session["current_header"] = item_form.cleaned_data.get(
                    "item_heading"
                )

                self.reload_cache(
                    headers=True,
                    header_items=["__all__"],
                )

                self.context["item_form"] = ItemForm(
                    initial={
                        "item_heading": self.request.session.get("current_header"),
                        "color": instance.color,
                    },
                    instance=instance,
                )
                self.context["format_reverse"] = True

                if (
                    DataItemSetModel.objects.filter(
                        user_code=self.request.session.get("user_code"),
                    ).exists()
                    and self.request.session.get("image_url", "") != ""
                ):
                    self.render_preview_url()

        # Loads Excel Sheet and Stores Data in Model
        elif (
            self.request.POST.get("submit") is not None
            and self.request.POST.get("submit") == "load_excel_submit"
        ):
            excel_form = ExcelForm(request.POST, request.FILES)

            if excel_form.is_valid():
                self.store_excel_to_model(self.request.FILES.get("excel_file"))

                self.context["excel_form"] = excel_form
                self.request.session["excel_file_name"] = self.request.FILES.get(
                    "excel_file"
                ).name
                self.context["excel_file_status"] = True

                self.reload_cache(
                    headers=True,
                    header_items=["__all__"],
                )

                self.context["inspector_header"] = cache.get(self.cache_key_header)
                self.request.session["current_header"] = cache.get(
                    self.cache_key_header
                )[0]

                if self.request.session.get("current_header", "") != "":
                    self.context["inspector_data"] = cache.get(
                        self.request.session.get("current_header")
                    )

                self.render_preview_url()

            else:
                self.context["excel_form"] = ExcelForm()
                self.request.session["excel_file_name"] = None
                self.context["excel_file_status"] = False
                self.context["excel_form_errors"]["has_error"] = True
                self.context["excel_form_errors"]["error"] = excel_form.errors

        # Load the Base Image for Output
        elif (
            self.request.POST.get("submit") is not None
            and self.request.POST.get("submit") == "load_image_submit"
        ):
            image_form = ImageForm(request.POST, request.FILES)

            if image_form.is_valid():
                try:
                    if self.request.session.get("image_file_name") is not None:
                        filter_instance = ImageModel.objects.filter(
                            user__unique_code=self.request.session.get("user_code"),
                        )

                    instance = ImageModel()

                    main_image_storage = ImageMediaLibrary("main")

                    image_upload_result = main_image_storage.upload_image(
                        image_file=self.request.FILES["image"],
                        tags=[
                            self.request.session.get("user_code"),
                            lambda verified_status: "verified"
                            if self.request.session.get("is_verified")
                            else "not verified",
                        ],
                        overwrite_status=True,
                    )

                    if image_upload_result:
                        instance.image_url = image_upload_result.url

                        instance.image_file_name = self.request.session[
                            "image_file_name"
                        ] = image_upload_result.name

                    else:
                        raise ImageMediaStorageError(
                            "Image Could not be Uploaded to the Media Storage"
                        )

                    with transaction.atomic():
                        instance.user = CustomUser.objects.get(
                            unique_code=self.request.session.get("user_code")
                        )

                        if filter_instance and filter_instance.exists():
                            filter_instance.delete()

                        instance.save()

                except Exception as e:
                    print("load_image" + e.__str__())
                    raise

                self.request.session["image_url"] = instance.image_url
                self.context["image_form"] = image_form
                self.context["image_status"] = True

                self.render_preview_url()

            else:
                self.context["image_form"] = ImageForm()
                self.request.session["image_file_name"] = None
                self.context["image_status"] = False
                self.request.session["image_url"] = None
                self.context["image_form_errors"]["has_errors"] = True
                self.context["image_form_errors"]["error"] = image_form.errors

        # Gets the Headers for Inspector Window
        elif self.request.POST.get("inspector_header_item", "") != "":
            self.request.session["current_header"] = self.request.POST.get(
                "inspector_header_item"
            )

            instance = DataItemSetModel.objects.filter(
                user_code=self.request.session.get("user_code")
            ).get(item_set_heading=self.request.session.get("current_header"))
            self.context["item_form"] = ItemForm(
                initial={
                    "item_heading": self.request.session.get("current_header"),
                    "color": instance.color,
                },
                instance=instance,
            )
            self.context["format_reverse"] = True

            self.context["inspector_data"] = cache.get(
                self.request.session.get("current_header")
            )

        # Remove a Header from the Inspector Window
        elif (
            self.request.POST.get("submit_remove") is not None
            and self.request.POST.get("submit_remove") == "inspector_header_item_remove"
        ):
            remove_header = self.request.POST.get("header_item")

            if self.request.session.get("user_code", "") != "":
                instance = DataItemSetModel.objects.get(
                    user_code=self.request.session.get("user_code"),
                    item_set_heading=remove_header,
                )

                if (
                    self.request.session.get("current_header")
                    == instance.item_set_heading
                ):
                    self.request.session["current_header"] = None

                instance.delete()

                cache.delete(instance.item_set_heading)

                self.reload_cache(headers=True, header_items=None)

            else:
                raise SessionValuesNotFoundError("user_code Not Available")

        # Updates the Data Set List in the Model
        elif (
            self.request.POST.get("submit") is not None
            and self.request.POST.get("submit") == "update_inspector_data"
        ):
            new_item_data = self.request.POST.getlist("inspector_data_item")

            if self.request.session.get("user_code", "") != "":
                try:
                    instance = DataItemSetModel.objects.get(
                        user_code=self.request.session.get("user_code"),
                        item_set_heading=self.request.session.get("current_header"),
                    )

                    instance.item_set = str(new_item_data)
                    instance.save()

                    if self.reload_cache(
                        headers=False,
                        header_items=[str(self.request.session.get("current_header"))],
                    ):
                        raise HeaderDataNotFoundError("Missing Header Item")

                    self.render_preview_url()

                except Exception as e:
                    print("update_inspector_data" + e.__str__())
                    raise

            else:
                raise SessionValuesNotFoundError("user_code Not Available")

        # Add the Inspector Data Item
        # use js to create an identical textbox with the name as inspector_data_item
        # and the value as the new value entered by the user
        # then prompt the user to click update changes

        # Remove the Inspector Data Item
        # use js to remove the respective textbox with the name as inspector_data_item
        # then prompt the user to click update changes

        # Export Images
        elif (
            self.request.POST.get("submit") is not None
            and self.request.POST.get("submit") == "export_images"
        ):
            export_form = ExportForm(request.POST)

            if export_form.is_valid():
                zip_path = render_and_zip_images(
                    request.session.get("user_code"), output_format="png"
                )

                # Create a response with the zip file
                response = HttpResponse(content_type="application/zip")
                response[
                    "Content-Disposition"
                ] = f'attachment; filename="autoficate_{request.session.get("user_name").split(request.session.get("user_code"))[0]}_output.zip"'

                # Open the zip file and write its content to the response
                with open(zip_path, "rb") as zip_file:
                    response.write(zip_file.read())

                # Clean up: Delete the zip file
                os.unlink(zip_path)

                # return response

        end_time = time.time()
        elapsed_time = end_time - start_time
        print(f"\nPOST took {elapsed_time:.6f} seconds to execute.\n")


def Custom404View(request, exception=None):
    return render(request, "Custom404.html", status=404)

@check_time
def SignupView(request):
    signup_form_errors = None
    db_errors = None

    if request.method == "POST":
        signup_form = SignUpForm(request.POST)

        if signup_form.is_valid():
            try:
                exists = False
                for user in CustomUser.objects.filter(
                    user_email__contains=signup_form.cleaned_data["user_email"]
                ).filter(user_email__endswith=".unregistered"):
                    DataItemSetModel.objects.filter(user_code=user.unique_code).delete()
                    exists = True

                if exists:
                    CustomUser.objects.filter(
                        user_email__contains=signup_form.cleaned_data["user_email"]
                    ).filter(user_email__endswith=".unregistered").delete()
                    exists = None

                if request.session.get("user_code", None) is not None:
                    instance = CustomUser.objects.get(
                        unique_code=request.session.get("user_code")
                    )

                    instance.user_email = signup_form.cleaned_data.get("user_email")
                    instance.update_password(
                        new_password=signup_form.cleaned_data.get("password1")
                    )

                    status, info = save_instance(instance)

                    if not status:
                        db_errors = info
                        render(
                            request,
                            "signup.html",
                            {
                                "signup_form": signup_form,
                                "signup_form_errors": db_errors,
                            },
                        )

                    user = authenticate(
                        request,
                        user_email=signup_form.cleaned_data.get("user_email"),
                        password=signup_form.cleaned_data.get("password1"),
                    )
                    login(request, user)
                    request.session["user_name"] = instance.username
                    request.session["user_code"] = instance.unique_code
                    request.session["is_verified"] = True

                else:
                    instance = signup_form.save()

                    group_instance = Group.objects.get(name="Users")
                    user_instance = CustomUser.objects.get(
                        unique_code=instance.unique_code
                    )
                    user_instance.groups.set([group_instance])

                    status, info = save_instance(user_instance)

                    if not status:
                        db_errors = info
                        render(
                            request,
                            "signup.html",
                            {
                                "signup_form": signup_form,
                                "signup_form_errors": db_errors,
                            },
                        )

                    user = authenticate(
                        request,
                        user_email=signup_form.cleaned_data.get("user_email"),
                        password=signup_form.cleaned_data.get("password1"),
                    )
                    login(request, user)
                    request.session["user_name"] = instance.username
                    request.session["user_code"] = instance.unique_code
                    request.session["is_verified"] = True

            except Exception as e:
                db_errors = e.__str__()
                render(
                    request,
                    "signup.html",
                    {"signup_form": signup_form, "signup_form_errors": db_errors},
                )

            return redirect("index")

        else:
            signup_form_errors = signup_form.errors

    else:
        if request.session.get("user_code", None) is not None:
            try:
                instance = CustomUser.objects.get(
                    unique_code=request.session.get("user_code")
                )

                signup_form = SignUpForm(
                    data={
                        "first_name": instance.first_name,
                        "last_name": instance.last_name,
                        "user_email": str(
                            instance.user_email.split(
                                f".{instance.unique_code}.unregistered"
                            )[0]
                        ),
                    }
                )

            except Exception as e:
                db_errors = e.__str__()
                signup_form = SignUpForm()

        else:
            signup_form = SignUpForm()

    return render(
        request,
        "signup.html",
        {"signup_form": signup_form, "signup_form_errors": signup_form_errors},
    )

@check_time
def LogoutView(request):
    cache.clear()
    logout(request)

    return redirect("index")


"""
def my_decorator(func):
    def wrapper(*args, **kwargs):
        instance = args[0]  # Assuming the first argument is the instance
        val = instance.a
        
        result = func(*args, **kwargs)

        print(result + val)
    
    return wrapper

##  Output : 8 ##

class c1:
    
    @my_decorator
    def my_function(self, a: int):
        self.a = a + 1
        return self.a

# Create an instance
obj = c1()

# Call the decorated function
obj.my_function(3)
"""
